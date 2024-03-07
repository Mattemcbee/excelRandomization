import openpyxl
import datetime

def filter_patients(file1_path, file2_path):
    wb1 = openpyxl.load_workbook(file1_path)
    sheet1 = wb1.active

    wb2 = openpyxl.load_workbook(file2_path)
    sheet2 = wb2.active

    result_wb = openpyxl.Workbook()
    result_sheet = result_wb.active

    drugName = 'drug'
    drug_col_name = 'Merged Cells'
    subject_col_name = 'Subject ID'

    decision_column_index_sheet1 = None
    for col_index, cell in enumerate(sheet1[1], start=1):
        if cell.value == drug_col_name:
            decision_column_index_sheet1 = col_index
            break

    if decision_column_index_sheet1 is None:
        print(f"Decision column not found ({drug_col_name}) in the first file.")
        return

    subject_column_index_sheet1 = None
    for col_index, cell in enumerate(sheet1[1], start=1):
        if cell.value == subject_col_name:
            subject_column_index_sheet1 = col_index
            break

    if subject_column_index_sheet1 is None:
        print(f"Subject column not found ({subject_col_name}) in the first file.")
        return

    subject_column_index_sheet2 = None
    for col_index, cell in enumerate(sheet2[1], start=1):
        if cell.value == subject_col_name:
            subject_column_index_sheet2 = col_index
            break

    if subject_column_index_sheet2 is None:
        print(f"Subject column not found ({subject_col_name}) in the second file.")
        return

    drug_subject_ids = set()

    for row in sheet1.iter_rows(min_row=2, values_only=False):
        subject_id = row[subject_column_index_sheet1 - 1].value
        decision = row[decision_column_index_sheet1 - 1].value

        if decision == drugName:
            drug_subject_ids.add(subject_id)

    for row2 in sheet2.iter_rows(min_row=2, values_only=False):
        subject_id = row2[subject_column_index_sheet2 - 1].value
        if subject_id in drug_subject_ids:
            result_row = [cell.value for cell in row2]
            result_sheet.append(result_row)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"pythonExcelResult_{timestamp}.xlsx"

    result_wb.save(output_path)
    print(f"Result saved to {output_path}")

# Example usage
filter_patients("pythonExcelNewRand.xlsx", "pythonExcelNewSub.xlsx")
